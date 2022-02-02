# importing openpyxl module
import openpyxl as xl;
from openpyxl import load_workbook
from openpyxl import Workbook

##
#Create Workbook
##
wb = Workbook()
ws =  wb.active
ws.title = "worksheet"
wb.save(filename = '/path/file.xlsx')


##
# opening the source excel file
##
wb1 = xl.load_workbook("/path/file1.xlsx")
ws1 = wb1.worksheets[0]

wb2 = xl.load_workbook("/path/file2.xlsx")
ws2 = wb2.worksheets[0]

wb3 = xl.load_workbook("/path/file3.xlsx")
ws3 = wb3.worksheets[0]
  
# opening the destination excel file 
filename1 ="/path/file.xlsx"
wb4 = xl.load_workbook(filename1)

wb4.create_sheet('Worksheet2')
wb4.create_sheet('Worksheet3')

ws4 = wb4.worksheets[0]
ws5 = wb4.worksheets[1]
ws6 = wb4.worksheets[2]
wb4.save('/path/sortieFinale.xlsx')




# calculate total number of rows and 
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column
  
# copying the cell values from source 
# excel file to destination excel file
for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws4.cell(row = i, column = j).value = c.value

# calculate total number of rows and 
# columns in source excel file
mr = ws2.max_row
mc = ws2.max_column
  
# copying the cell values from source 
# excel file to destination excel file
for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        # reading cell value from source excel file
        c = ws2.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws5.cell(row = i, column = j).value = c.value


# calculate total number of rows and 
# columns in source excel file
mr = ws3.max_row
mc = ws3.max_column
  
# copying the cell values from source 
# excel file to destination excel file
for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        # reading cell value from source excel file
        c = ws3.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws6.cell(row = i, column = j).value = c.value


  
# saving the destination excel file
wb4.save(str(filename1))


